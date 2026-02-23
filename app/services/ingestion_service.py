import importlib
import logging
import re
import threading
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select as sa_select
from sqlalchemy.exc import SQLAlchemyError

from app.config import get_settings
from app.core.constants import STATIC_DIR
from app.database import Base, SessionLocal, engine, ensure_sqlite_schema
from app.models.inventory import Inventory
from app.models.product import Product
from app.models.stores import Store
from app.services.product_service import apply_price_update

logger = logging.getLogger(__name__)

_IMAGE_DIR = STATIC_DIR / "images"
_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp")
_IMAGE_INDEX_CACHE: dict[str, object] = {"mtime": None, "index": {}}


def _find_header_index(header_keys, target):
    for idx, key in enumerate(header_keys):
        if key == target:
            return idx
    return None


def _sanitize_image_basename(value):
    if not value:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    sanitized = []
    for char in text:
        if char.isalnum() or char in ("-", "_", "."):
            sanitized.append(char)
        else:
            sanitized.append("_")
    return "".join(sanitized).strip("_.")


def _image_extension_from_format(image_format):
    if not image_format:
        return ".jpg"
    fmt = str(image_format).strip().lower()
    if fmt in {"jpg", "jpeg"}:
        return ".jpg"
    if fmt in {"png", "webp"}:
        return f".{fmt}"
    return ".jpg"


def _looks_like_explicit_image_reference(value):
    if _is_blank(value):
        return False
    cleaned = str(value).strip().replace("\\", "/")
    if not cleaned:
        return False
    lower_value = cleaned.lower()
    if lower_value.startswith(("http://", "https://", "data:", "/")):
        return True
    if lower_value.startswith(("static/", "images/")):
        return True
    return Path(cleaned).suffix.lower() in _IMAGE_EXTENSIONS


def _extract_embedded_images(worksheet, header_keys):
    image_col = _find_header_index(header_keys, "image_url")
    if image_col is None:
        return {}
    style_col = _find_header_index(header_keys, "style_code")
    images = getattr(worksheet, "_images", []) or []
    if not images:
        return {}
    _IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    image_map = {}
    for image in images:
        anchor = getattr(image, "anchor", None)
        if not anchor or not hasattr(anchor, "_from"):
            continue
        cell_from = anchor._from
        col = getattr(cell_from, "col", None)
        row = getattr(cell_from, "row", None)
        if col is None or row is None:
            continue
        if col != image_col:
            continue
        row_idx = row + 1
        style_value = None
        if style_col is not None:
            style_value = worksheet.cell(row=row_idx, column=style_col + 1).value
        base_name = _sanitize_image_basename(style_value) or f"row_{row_idx}"
        extension = _image_extension_from_format(getattr(image, "format", None))
        filename = f"{base_name}{extension}"
        try:
            data = image._data()
        except Exception:
            data = None
        if not data:
            continue
        try:
            (_IMAGE_DIR / filename).write_bytes(data)
        except OSError:
            continue
        image_map[row_idx] = f"/static/images/{filename}"
    return image_map

_ALIAS_SPECS = (
    (("store", "id"), "store_id"),
    (("product", "id"), "product_id"),
    (("style", "code"), "style_code"),
    (("article", "name"), "article_name"),
    (("image", "url"), "image_url"),
    (("image", "link"), "image_url"),
    (("image",), "image_url"),
    (("supplier", "name"), "supplier_name"),
    (("department", "name"), "department_name"),
    (("category", "name"), "category"),
    (("item", "mrp"), "mrp"),
    (("price",), "price"),
    (("stock", "days"), "stock_days"),
    (("cbs", "qty"), "quantity"),
    (("qty",), "quantity"),
    (("cost", "price"), "cost_price"),
    (("current", "price"), "current_price"),
    (("lifecycle", "start"), "lifecycle_start_date"),
    (("lifecycle", "start", "date"), "lifecycle_start_date"),
    (("start", "date"), "lifecycle_start_date"),
)

HEADER_ALIASES = {"".join(parts): target for parts, target in _ALIAS_SPECS}

REQUIRED_COLUMNS = {
    "stores": {"name", "city"},
    "products": {
        "store_id",
        "style_code",
        "barcode",
        "article_name",
        "category",
        "supplier_name",
        "mrp",
    },
    "inventory": {
        "store_id",
        "product_id",
        "quantity",
        "cost_price",
        "current_price",
        "lifecycle_start_date",
    },
    "daily_update": {
        "store_id",
        "style_code",
        "supplier_name",
        "stock_days",
        "department_name",
        "category",
        "mrp",
    },
}

DAILY_UPDATE_SHEET = "daily_update"
DEFAULT_SHEET_ORDER = ["stores", "products", "inventory"]
CARD_LAYOUT_LABEL_MAP = {
    "supplier_name": "supplier_name",
    "supplier": "supplier_name",
    "department": "department_name",
    "department_name": "department_name",
    "style": "style_code",
    "style_code": "style_code",
    "itemcode": "barcode",
    "item_code": "barcode",
    "barcode": "barcode",
    "mrp": "mrp",
    "item_mrp": "mrp",
    "image_name": "image_url",
    "image_url": "image_url",
    "image": "image_url",
    "pur_qty": "quantity",
    "qty": "quantity",
    "quantity": "quantity",
    "cbs_qty": "quantity",
    "stock_days": "stock_days",
    "category_name": "category",
    "category": "category",
    "store_id": "store_id",
}
_PLACEHOLDER_VALUES = {"none", "[none]", "null", "[null]", "na", "n/a", "nan", "-", "--"}
_LEADING_STORE_CODE = re.compile(r"^\s*(\d{3})")


def _import_models():
    for module_name in (
        "app.models.alert",
        "app.models.daily_snapshot",
        "app.models.delivery_logs",
        "app.models.inventory",
        "app.models.lifecycle",
        "app.models.price_history",
        "app.models.product",
        "app.models.risk_log",
        "app.models.sales",
        "app.models.stores",
    ):
        importlib.import_module(module_name)


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _clean_text(value):
    if _is_blank(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.lower() in _PLACEHOLDER_VALUES:
        return None
    return text


def _is_summary_value(value):
    if isinstance(value, str):
        value_text = value.strip().lower()
        if value_text and "total" in value_text:
            return True
    return False


def _is_footer_value(value):
    if not isinstance(value, str):
        return False
    value_text = value.strip().lower()
    return value_text.startswith("printed on") or value_text.startswith("generated on")


def _looks_like_header_row(row, header_set):
    matches = 0
    non_blank = 0
    for value in row:
        if _is_blank(value):
            continue
        non_blank += 1
        if normalize_header(value) in header_set:
            matches += 1
    return 0 < non_blank == matches


def normalize_header(value):
    if value is None:
        return ""
    value_text = str(value).strip().lower()
    if not value_text:
        return ""
    for char in (" ", "-", ".", "/"):
        value_text = value_text.replace(char, "_")
    value_text = "_".join(part for part in value_text.split("_") if part)
    alias = HEADER_ALIASES.get(value_text)
    if alias:
        return alias
    alias = HEADER_ALIASES.get(value_text.replace("_", ""))
    if alias:
        return alias
    return value_text


def normalize_sheet_name(name):
    value_text = str(name).strip().lower()
    if not value_text:
        return ""
    for char in (" ", "-", ".", "/"):
        value_text = value_text.replace(char, "_")
    return "_".join(part for part in value_text.split("_") if part)


def normalize_sheet_list(value):
    if not value:
        return None
    if isinstance(value, (list, tuple, set)):
        items = [str(item).strip() for item in value if str(item).strip()]
    else:
        items = [part.strip() for part in str(value).split(",") if part.strip()]
    return items or None


def _infer_store_id_from_sheet_name(sheet_name):
    if _is_blank(sheet_name):
        return None
    match = _LEADING_STORE_CODE.match(str(sheet_name))
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _to_card_field(label):
    normalized = normalize_header(label)
    return CARD_LAYOUT_LABEL_MAP.get(normalized)


def _build_daily_update_row_from_card(record, inferred_store_id):
    supplier_name = _clean_text(record.get("supplier_name"))
    if not supplier_name:
        return None

    style_code = _clean_text(record.get("style_code"))
    barcode = _clean_text(record.get("barcode"))
    if not style_code:
        style_code = barcode
    if not style_code:
        return None

    mrp = record.get("mrp")
    if _is_blank(mrp):
        return None

    store_value = record.get("store_id")
    store_id = None
    if not _is_blank(store_value):
        store_id = to_int(store_value, "store_id", required=False)
    if store_id is None:
        store_id = inferred_store_id
    if store_id is None:
        return None

    department_name = _clean_text(record.get("department_name"))
    category = _clean_text(record.get("category"))
    if not department_name:
        department_name = category or "Uncategorized"
    if not category:
        category = department_name

    stock_days = record.get("stock_days")
    if _is_blank(stock_days):
        stock_days = 0

    row = {
        "store_id": store_id,
        "supplier_name": supplier_name,
        "stock_days": stock_days,
        "style_code": style_code,
        "department_name": department_name,
        "category": category,
        "mrp": mrp,
    }

    quantity = record.get("quantity")
    if not _is_blank(quantity):
        row["quantity"] = quantity

    image_url = _clean_text(record.get("image_url"))
    if image_url:
        row["image_url"] = image_url

    if barcode:
        row["barcode"] = barcode

    return row


def _append_card_record(rows, columns, record_metas, record, inferred_store_id):
    if not record:
        return
    converted = _build_daily_update_row_from_card(record, inferred_store_id)
    if not converted:
        return
    rows.append(converted)
    columns.update(converted.keys())
    value_idx = record.get("__value_idx")
    supplier_row = record.get("__supplier_row")
    if isinstance(value_idx, int) and isinstance(supplier_row, int):
        record_metas.append((len(rows) - 1, value_idx, supplier_row))


def _extract_card_layout_images(worksheet, rows, record_metas):
    images = getattr(worksheet, "_images", []) or []
    if not images or not rows or not record_metas:
        return

    value_indices = sorted({meta[1] for meta in record_metas})
    if not value_indices:
        return

    records_by_value_idx = {}
    for row_index, value_idx, supplier_row in record_metas:
        records_by_value_idx.setdefault(value_idx, []).append((supplier_row, row_index))
    for value_idx in records_by_value_idx:
        records_by_value_idx[value_idx].sort(key=lambda item: item[0])

    anchored_images = []
    for image in images:
        anchor = getattr(image, "anchor", None)
        if not anchor or not hasattr(anchor, "_from"):
            continue
        cell_from = anchor._from
        col = getattr(cell_from, "col", None)
        row = getattr(cell_from, "row", None)
        if col is None or row is None:
            continue
        anchored_images.append((row + 1, col, image))

    if not anchored_images:
        return

    anchored_images.sort(key=lambda item: (item[0], item[1]))
    used_rows = set()
    _IMAGE_DIR.mkdir(parents=True, exist_ok=True)

    for image_row, image_col, image in anchored_images:
        right_side_indices = [idx for idx in value_indices if idx > image_col]
        if right_side_indices:
            value_idx = min(right_side_indices, key=lambda idx: idx - image_col)
        else:
            value_idx = min(value_indices, key=lambda idx: abs(idx - image_col))

        candidates = records_by_value_idx.get(value_idx) or []
        if not candidates:
            continue

        supplier_target = image_row + 1
        best_row_index = None
        best_distance = None
        for supplier_row, row_index in candidates:
            if row_index in used_rows:
                continue
            distance = abs(supplier_row - supplier_target)
            if best_distance is None or distance < best_distance:
                best_distance = distance
                best_row_index = row_index
        if best_row_index is None:
            continue

        try:
            data = image._data()
        except Exception:
            data = None
        if not data:
            continue

        record = rows[best_row_index]
        base_value = (
            _clean_text(record.get("image_url"))
            or _clean_text(record.get("style_code"))
            or _clean_text(record.get("barcode"))
            or f"row_{best_row_index + 1}"
        )
        base_name = _sanitize_image_basename(base_value) or f"row_{best_row_index + 1}"
        extension = _image_extension_from_format(getattr(image, "format", None))
        filename = f"{base_name}{extension}"
        try:
            (_IMAGE_DIR / filename).write_bytes(data)
        except OSError:
            continue
        record["image_url"] = f"/static/images/{filename}"
        used_rows.add(best_row_index)


def load_card_layout_rows(worksheet):
    rows = []
    columns = set()
    record_metas = []
    active_records = {}
    inferred_store_id = _infer_store_id_from_sheet_name(worksheet.title)

    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for col_idx, raw_label in enumerate(row):
            key = _to_card_field(raw_label)
            if not key:
                continue
            value_idx = col_idx + 2
            if value_idx >= len(row):
                continue
            value = row[value_idx]
            if _is_blank(value):
                continue

            record = active_records.get(value_idx)
            if key == "supplier_name":
                _append_card_record(rows, columns, record_metas, record, inferred_store_id)
                record = {"__value_idx": value_idx, "__supplier_row": row_number}
                active_records[value_idx] = record
            elif record is None:
                record = {"__value_idx": value_idx, "__supplier_row": row_number}
                active_records[value_idx] = record

            record[key] = value

    for record in active_records.values():
        _append_card_record(rows, columns, record_metas, record, inferred_store_id)

    _extract_card_layout_images(worksheet, rows, record_metas)

    return rows, columns


def get_daily_update_aliases():
    settings = get_settings()
    aliases = normalize_sheet_list(settings.EXCEL_DAILY_UPDATE_SHEET_ALIASES)
    if not aliases:
        return []
    return [normalize_sheet_name(name) for name in aliases]


def apply_daily_update_aliases(sheet_map, aliases):
    if not aliases or DAILY_UPDATE_SHEET in sheet_map:
        return
    for alias in aliases:
        actual_name = sheet_map.get(alias)
        if actual_name:
            sheet_map[DAILY_UPDATE_SHEET] = actual_name
            break


def should_create_missing_stores():
    settings = get_settings()
    return bool(settings.EXCEL_CREATE_MISSING_STORES)


def ensure_store_exists(db, store_id):
    if not should_create_missing_stores():
        return
    if store_id is None:
        return
    if db.get(Store, store_id):
        return
    db.add(Store(id=store_id, name=f"Store {store_id}", city="Unknown"))
    db.flush()


def to_str(value, field, required=True):
    if _is_blank(value):
        if required:
            raise ValueError(f"{field} is required")
        return None
    return str(value).strip()


def to_int(value, field, required=True):
    if _is_blank(value):
        if required:
            raise ValueError(f"{field} is required")
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field} must be an integer")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"{field} must be an integer")
    if isinstance(value, str):
        value_text = value.strip()
        try:
            return int(value_text)
        except ValueError:
            try:
                numeric = float(value_text)
            except ValueError:
                raise ValueError(f"{field} must be an integer") from None
            if not numeric.is_integer():
                raise ValueError(f"{field} must be an integer")
            return int(numeric)
    return int(value)


def to_float(value, field):
    if _is_blank(value):
        raise ValueError(f"{field} is required")
    if isinstance(value, str):
        value = value.replace(",", "")
    return float(value)


def to_date(value, field):
    if _is_blank(value):
        raise ValueError(f"{field} is required")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value_text = value.strip()
        try:
            return date.fromisoformat(value_text)
        except ValueError:
            pass
        for fmt in ("%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value_text, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"{field} must be a date (YYYY-MM-DD)")


def load_sheet_rows(worksheet):
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return [], set()
    header_keys = [normalize_header(header) for header in headers]
    non_blank_headers = [key for key in header_keys if key]
    should_try_card_layout = False
    if not non_blank_headers:
        should_try_card_layout = True
    elif len(non_blank_headers) < 3:
        should_try_card_layout = True
    elif len(set(non_blank_headers)) <= len(non_blank_headers) // 2:
        should_try_card_layout = True

    if should_try_card_layout:
        card_rows, card_columns = load_card_layout_rows(worksheet)
        if card_rows:
            return card_rows, card_columns
        if not non_blank_headers:
            return [], set()
    indices = [(idx, key) for idx, key in enumerate(header_keys) if key]
    columns = {key for key in header_keys if key}
    embedded_images = _extract_embedded_images(worksheet, header_keys)
    store_idx = None
    for idx, key in indices:
        if key == "store_id":
            store_idx = idx
            break

    rows = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if row is None or all(_is_blank(value) for value in row):
            continue
        if _looks_like_header_row(row, columns):
            continue
        if store_idx is not None and store_idx < len(row):
            store_value = row[store_idx]
            if _is_blank(store_value):
                continue
            if _is_summary_value(store_value):
                continue
            if _is_footer_value(store_value):
                continue
        record = {key: row[idx] for idx, key in indices}
        if embedded_images:
            image_value = embedded_images.get(row_idx)
            if image_value and (
                _is_blank(record.get("image_url"))
                or not _looks_like_explicit_image_reference(record.get("image_url"))
            ):
                record["image_url"] = image_value
        rows.append(record)
    return rows, columns


def validate_columns(sheet_name, columns):
    required = REQUIRED_COLUMNS[sheet_name]
    missing = sorted(required - columns)
    if sheet_name == DAILY_UPDATE_SHEET and "stock_days" in missing:
        if "lifecycle_start_date" in columns:
            missing.remove("stock_days")
    if sheet_name == DAILY_UPDATE_SHEET and "category" in missing:
        if "department_name" in columns:
            missing.remove("category")
    if sheet_name == "inventory" and "cost_price" in missing:
        if "mrp" in columns:
            missing.remove("cost_price")
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(f"{sheet_name} sheet missing columns: {missing_text}")


def get_existing(db, model, record_id, *filters):
    if record_id:
        instance = db.get(model, record_id)
        if instance:
            return instance
    if filters:
        return db.execute(sa_select(model).where(*filters)).scalars().first()
    return None


def apply_upsert(db, instance, model, values):
    if instance:
        for key, value in values.items():
            setattr(instance, key, value)
        return "updated"
    db.add(model(**values))
    return "inserted"


def upsert_store(db, row):
    store_id = to_int(row.get("id"), "id", required=False)
    name = to_str(row.get("name"), "name")
    city = to_str(row.get("city"), "city")

    store = get_existing(db, Store, store_id, Store.name == name, Store.city == city)
    values = {"name": name, "city": city}
    return apply_upsert(db, store, Store, values)


def upsert_product(db, row, *, price_change_log=None):
    product_id = to_int(row.get("id"), "id", required=False)
    store_id = to_int(row.get("store_id"), "store_id")
    ensure_store_exists(db, store_id)
    style_code = to_str(row.get("style_code"), "style_code")
    barcode = to_str(row.get("barcode"), "barcode")
    article_name = to_str(row.get("article_name"), "article_name")
    category = to_str(row.get("category"), "category")
    supplier_name = to_str(row.get("supplier_name"), "supplier_name")
    mrp = to_float(row.get("mrp"), "mrp")
    price = parse_optional_price(row)
    image_url, image_explicit = resolve_image_url(row)
    department_value = row.get("department_name")
    department_name = None
    if not _is_blank(department_value):
        department_name = to_str(department_value, "department_name")

    product = get_existing(
        db,
        Product,
        product_id,
        Product.style_code == style_code,
        Product.store_id == store_id,
    )
    values = build_product_values(
        store_id=store_id,
        style_code=style_code,
        barcode=barcode,
        article_name=article_name,
        category=category,
        supplier_name=supplier_name,
        mrp=mrp,
        product=product,
        image_url=image_url,
        image_explicit=image_explicit,
        department_name=department_name,
        default_department_for_new=True,
    )
    return finalize_product_upsert(
        db,
        product,
        values,
        mrp=mrp,
        price=price,
        store_id=store_id,
        style_code=style_code,
        price_change_log=price_change_log,
    )


def upsert_inventory(db, row):
    inventory_id = to_int(row.get("id"), "id", required=False)
    store_id = to_int(row.get("store_id"), "store_id")
    ensure_store_exists(db, store_id)
    product_id = to_int(row.get("product_id"), "product_id")
    quantity = to_int(row.get("quantity"), "quantity")
    cost_price = resolve_price(row.get("cost_price"), row.get("mrp"), "item_mrp")
    current_price = to_float(row.get("current_price"), "current_price")
    lifecycle_start_date = to_date(row.get("lifecycle_start_date"), "lifecycle_start_date")

    inventory = get_existing(
        db,
        Inventory,
        inventory_id,
        Inventory.store_id == store_id,
        Inventory.product_id == product_id,
    )
    values = {
        "store_id": store_id,
        "product_id": product_id,
        "quantity": quantity,
        "cost_price": cost_price,
        "current_price": current_price,
        "lifecycle_start_date": lifecycle_start_date,
    }
    return apply_upsert(db, inventory, Inventory, values)


def resolve_price(value, fallback, field):
    if not _is_blank(value):
        return to_float(value, field)
    if _is_blank(fallback):
        raise ValueError(f"{field} is required")
    return to_float(fallback, field)


def resolve_lifecycle_start_date(row):
    lifecycle_value = row.get("lifecycle_start_date")
    if not _is_blank(lifecycle_value):
        return to_date(lifecycle_value, "lifecycle_start_date")
    stock_days = to_int(row.get("stock_days"), "stock_days")
    return date.today() - timedelta(days=stock_days)


def parse_optional_price(row):
    price_value = row.get("price")
    if not _is_blank(price_value):
        return to_float(price_value, "price")
    return None


def _build_image_index():
    if not _IMAGE_DIR.exists():
        return {}
    index: dict[str, str] = {}
    for path in _IMAGE_DIR.iterdir():
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        if ext not in _IMAGE_EXTENSIONS:
            continue
        filename = path.name
        index[filename.lower()] = filename
        index[path.stem.lower()] = filename
    return index


def _get_image_index():
    if not _IMAGE_DIR.exists():
        return {}
    try:
        mtime = _IMAGE_DIR.stat().st_mtime_ns
    except OSError:
        return {}
    cached_mtime = _IMAGE_INDEX_CACHE.get("mtime")
    if cached_mtime != mtime:
        _IMAGE_INDEX_CACHE["index"] = _build_image_index()
        _IMAGE_INDEX_CACHE["mtime"] = mtime
    return _IMAGE_INDEX_CACHE.get("index") or {}


def _normalize_image_value(value, image_index):
    if _is_blank(value):
        return None
    cleaned = to_str(value, "image_url", required=False)
    if not cleaned:
        return None
    cleaned = cleaned.replace("\\", "/")
    lower_value = cleaned.lower()
    if lower_value.startswith(("http://", "https://", "data:", "/")):
        return cleaned
    if lower_value.startswith("static/"):
        return f"/{cleaned}"
    if lower_value.startswith("images/"):
        return f"/static/{cleaned}"
    filename = image_index.get(lower_value)
    if not filename:
        filename = image_index.get(Path(cleaned).stem.lower())
    if filename:
        return f"/static/images/{filename}"
    if Path(cleaned).suffix.lower() in _IMAGE_EXTENSIONS:
        return f"/static/images/{cleaned}"
    return None


def resolve_image_url(row):
    image_index = _get_image_index()
    explicit_value = row.get("image_url")
    if not _is_blank(explicit_value):
        return _normalize_image_value(explicit_value, image_index), True
    for key in ("style_code", "barcode"):
        candidate = row.get(key)
        if _is_blank(candidate):
            continue
        lookup_key = str(candidate).strip().lower()
        filename = image_index.get(lookup_key)
        if filename:
            return f"/static/images/{filename}", False
    return None, False


def warn_store_mismatch(product, store_id, style_code):
    if product.store_id != store_id:
        logger.warning(
            "Style code %s already exists for store %s; received store %s.",
            style_code,
            product.store_id,
            store_id,
        )


def apply_product_updates(product, values):
    for key, value in values.items():
        if key == "store_id":
            continue
        setattr(product, key, value)


def build_product_values(
    *,
    store_id,
    style_code,
    barcode,
    article_name,
    category,
    supplier_name,
    mrp,
    product,
    image_url,
    image_explicit,
    department_name=None,
    default_department_for_new=False,
):
    values = {
        "store_id": store_id,
        "style_code": style_code,
        "barcode": barcode,
        "article_name": article_name,
        "category": category,
        "supplier_name": supplier_name,
        "mrp": mrp,
    }
    if image_explicit:
        values["image_url"] = image_url
    elif image_url is not None and (not product or not product.image_url):
        values["image_url"] = image_url
    if department_name is not None:
        values["department_name"] = department_name
    elif default_department_for_new and not product:
        values["department_name"] = ""
    return values


def finalize_product_upsert(
    db,
    product,
    values,
    *,
    mrp,
    price,
    store_id,
    style_code,
    return_product=False,
    flush_on_insert=False,
    price_change_log=None,
):
    if product:
        old_price = product.price
        old_mrp = product.mrp
        warn_store_mismatch(product, store_id, style_code)
        apply_product_updates(product, values)
        changed_at = None
        if price is not None:
            changed_at = apply_price_update(db, product, price)
            if changed_at and price_change_log is not None:
                price_change_log.append(
                    {
                        "store_id": store_id,
                        "style_code": style_code,
                        "old_price": old_price if old_price is not None else 0.0,
                        "new_price": price,
                        "changed_at": changed_at.isoformat(),
                        "source": "price",
                    }
                )
        else:
            if (
                mrp is not None
                and old_mrp is not None
                and float(mrp) != float(old_mrp)
                and (old_price is None or float(old_price) == float(old_mrp))
            ):
                changed_at = apply_price_update(db, product, mrp)
                if changed_at and price_change_log is not None:
                    price_change_log.append(
                        {
                            "store_id": store_id,
                            "style_code": style_code,
                            "old_price": old_price if old_price is not None else 0.0,
                            "new_price": mrp,
                            "changed_at": changed_at.isoformat(),
                            "source": "mrp",
                        }
                    )
        if return_product:
            return "updated", product
        return "updated"

    product = build_product(values, mrp, price)
    db.add(product)
    if flush_on_insert:
        db.flush()
    if return_product:
        return "inserted", product
    return "inserted"


def build_product(values, mrp, price):
    if price is None:
        price = mrp
    product = Product(**values, price=price)
    if price is not None:
        product.last_price_update = datetime.now(timezone.utc)
    return product


def upsert_product_from_daily_update(db, row, *, price_change_log=None):
    store_id = to_int(row.get("store_id"), "store_id")
    ensure_store_exists(db, store_id)
    style_code = to_str(row.get("style_code"), "style_code")
    barcode_value = row.get("barcode")
    barcode = (
        to_str(barcode_value, "barcode")
        if not _is_blank(barcode_value)
        else style_code
    )
    article_value = row.get("article_name")
    article_name = (
        to_str(article_value, "article_name")
        if not _is_blank(article_value)
        else style_code
    )
    category_value = row.get("category")
    if _is_blank(category_value):
        category_value = row.get("department_name")
    category = to_str(category_value, "category")
    supplier_name = to_str(row.get("supplier_name"), "supplier_name")
    mrp = to_float(row.get("mrp"), "mrp")
    price = parse_optional_price(row)
    image_url, image_explicit = resolve_image_url(row)
    department_name = to_str(row.get("department_name"), "department_name")

    product = get_existing(
        db,
        Product,
        None,
        Product.style_code == style_code,
        Product.store_id == store_id,
    )
    values = build_product_values(
        store_id=store_id,
        style_code=style_code,
        barcode=barcode,
        article_name=article_name,
        category=category,
        supplier_name=supplier_name,
        mrp=mrp,
        product=product,
        image_url=image_url,
        image_explicit=image_explicit,
        department_name=department_name,
    )
    return finalize_product_upsert(
        db,
        product,
        values,
        mrp=mrp,
        price=price,
        store_id=store_id,
        style_code=style_code,
        return_product=True,
        flush_on_insert=True,
        price_change_log=price_change_log,
    )


def upsert_inventory_from_daily_update(db, row, product):
    store_id = product.store_id
    product_id = product.id
    quantity = to_int(row.get("quantity"), "quantity", required=False)
    if quantity is None:
        quantity = 0
    mrp = to_float(row.get("mrp"), "mrp")
    cost_price = resolve_price(row.get("cost_price"), mrp, "cost_price")
    current_price = resolve_price(row.get("current_price"), mrp, "current_price")
    lifecycle_start_date = resolve_lifecycle_start_date(row)

    inventory = get_existing(
        db,
        Inventory,
        None,
        Inventory.store_id == store_id,
        Inventory.product_id == product_id,
    )
    values = {
        "store_id": store_id,
        "product_id": product_id,
        "quantity": quantity,
        "cost_price": cost_price,
        "current_price": current_price,
        "lifecycle_start_date": lifecycle_start_date,
    }
    return apply_upsert(db, inventory, Inventory, values)


def import_daily_update_row(db, row, *, price_change_log=None):
    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    product_action, product = upsert_product_from_daily_update(
        db, row, price_change_log=price_change_log
    )
    counts[product_action] += 1
    inventory_action = upsert_inventory_from_daily_update(db, row, product)
    counts[inventory_action] += 1
    return counts


def import_rows(db, sheet_name, rows):
    counts = {"inserted": 0, "updated": 0, "skipped": 0, "price_changes": []}
    price_change_log = counts["price_changes"]
    for row in rows:
        if sheet_name == DAILY_UPDATE_SHEET:
            row_counts = import_daily_update_row(db, row, price_change_log=price_change_log)
            for key, value in row_counts.items():
                counts[key] += value
            continue
        if sheet_name == "stores":
            action = upsert_store(db, row)
        elif sheet_name == "products":
            action = upsert_product(db, row, price_change_log=price_change_log)
        elif sheet_name == "inventory":
            action = upsert_inventory(db, row)
        else:
            raise ValueError(f"Unsupported sheet: {sheet_name}")
        counts[action] += 1
    return counts


def import_workbook(workbook_path, sheets=None, dry_run=False):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"File not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported.")

    workbook = load_workbook(workbook_path, data_only=True)
    sheet_map = {normalize_sheet_name(name): name for name in workbook.sheetnames}
    daily_update_aliases = get_daily_update_aliases()
    apply_daily_update_aliases(sheet_map, daily_update_aliases)

    sheet_list = normalize_sheet_list(sheets)
    if sheet_list:
        requested = []
        for name in sheet_list:
            key = normalize_sheet_name(name)
            if key in daily_update_aliases:
                key = DAILY_UPDATE_SHEET
            requested.append(key)
    else:
        if DAILY_UPDATE_SHEET in sheet_map:
            requested = [DAILY_UPDATE_SHEET]
        else:
            requested = [name for name in DEFAULT_SHEET_ORDER if name in sheet_map]

    if not requested:
        if len(sheet_map) == 1:
            only_key = next(iter(sheet_map))
            sheet_map[DAILY_UPDATE_SHEET] = sheet_map[only_key]
            requested = [DAILY_UPDATE_SHEET]
        else:
            raise ValueError("No matching sheets found to import.")

    _import_models()
    Base.metadata.create_all(bind=engine)
    ensure_sqlite_schema()

    results = {}
    db = SessionLocal()
    try:
        for sheet_key in requested:
            actual_name = sheet_map.get(sheet_key)
            if not actual_name:
                raise ValueError(f"Sheet not found: {sheet_key}")
            worksheet = workbook[actual_name]
            rows, columns = load_sheet_rows(worksheet)
            validate_columns(sheet_key, columns)
            results[sheet_key] = import_rows(db, sheet_key, rows)

        if dry_run:
            db.rollback()
        else:
            db.commit()
    except SQLAlchemyError:
        db.rollback()
        raise
    finally:
        db.close()

    return results


def ensure_datasource_dir(path):
    datasource_dir = Path(path)
    datasource_dir.mkdir(parents=True, exist_ok=True)
    return datasource_dir


def summarize_results(results):
    parts = []
    for sheet_name, counts in results.items():
        parts.append(
            "{}: {} inserted, {} updated".format(
                sheet_name,
                counts.get("inserted", 0),
                counts.get("updated", 0),
            )
        )
    return "; ".join(parts) if parts else "no rows"


class ExcelWatchService:
    def __init__(self, watch_dir, poll_seconds=10, sheets=None):
        self.watch_dir = Path(watch_dir)
        self.poll_seconds = max(2, int(poll_seconds))
        self.sheets = normalize_sheet_list(sheets)
        self._stop_event = threading.Event()
        self._thread = None
        self._lock = threading.Lock()
        self._seen = {}
        self._processed = {}

    def start(self):
        if self._thread and self._thread.is_alive():
            return
        ensure_datasource_dir(self.watch_dir)
        self._stop_event.clear()
        self._thread = threading.Thread(
            target=self._run,
            name="excel-auto-import",
            daemon=True,
        )
        self._thread.start()
        logger.info("Excel auto-import watching: %s", self.watch_dir)

    def stop(self):
        if not self._thread:
            return
        self._stop_event.set()
        self._thread.join(timeout=self.poll_seconds + 1)
        self._thread = None
        logger.info("Excel auto-import stopped")

    def _run(self):
        while not self._stop_event.is_set():
            self._scan_once()
            self._stop_event.wait(self.poll_seconds)

    def _scan_once(self):
        if not self.watch_dir.exists():
            return
        for file_path in sorted(self.watch_dir.glob("*.xlsx")):
            if not self._is_candidate(file_path):
                continue
            if not self._is_ready(file_path):
                continue
            self._import_file(file_path)

    @staticmethod
    def _is_candidate(file_path):
        if not file_path.is_file():
            return False
        if file_path.name.startswith("~$"):
            return False
        return True

    def _is_ready(self, file_path):
        try:
            stat = file_path.stat()
        except OSError:
            return False
        key = (stat.st_mtime, stat.st_size)
        last_seen = self._seen.get(file_path)
        self._seen[file_path] = key
        if last_seen != key:
            return False
        last_processed = self._processed.get(file_path)
        if last_processed is None or stat.st_mtime > last_processed:
            return True
        return False

    def _import_file(self, file_path):
        with self._lock:
            try:
                stat = file_path.stat()
            except OSError:
                return
            try:
                results = import_workbook(file_path, sheets=self.sheets, dry_run=False)
            except (OSError, ValueError, SQLAlchemyError, InvalidFileException) as exc:
                logger.exception("Excel import failed for %s: %s", file_path, exc)
                return
            self._processed[file_path] = stat.st_mtime
            logger.info(
                "Excel import completed for %s: %s",
                file_path,
                summarize_results(results),
            )

import importlib
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select as sa_select
from sqlalchemy.exc import SQLAlchemyError

from app.config import get_settings
from app.database import Base, SessionLocal, engine, ensure_sqlite_schema
from app.models.inventory import Inventory
from app.models.product import Product
from app.models.stores import Store

_ALIAS_SPECS = (
    (("store", "id"), "store_id"),
    (("product", "id"), "product_id"),
    (("style", "code"), "style_code"),
    (("article", "name"), "article_name"),
    (("supplier", "name"), "supplier_name"),
    (("department", "name"), "department_name"),
    (("category", "name"), "category"),
    (("item", "mrp"), "mrp"),
    (("stock", "days"), "stock_days"),
    (("cbs", "qty"), "quantity"),
    (("qty",), "quantity"),
    (("cost", "price"), "cost_price"),
    (("current", "price"), "current_price"),
    (("lifecycle", "start"), "lifecycle_start_date"),
    (("lifecycle", "start", "date"), "lifecycle_start_date"),
    (("start", "date"), "lifecycle_start_date"),
)

HEADER_ALIASES = {"".join(parts): target for parts, target in _ALIAS_SPECS}

REQUIRED_COLUMNS = {
    "stores": {"name", "city"},
    "products": {
        "store_id",
        "style_code",
        "barcode",
        "article_name",
        "category",
        "supplier_name",
        "mrp",
    },
    "inventory": {
        "store_id",
        "product_id",
        "quantity",
        "cost_price",
        "current_price",
        "lifecycle_start_date",
    },
    "daily_update": {
        "store_id",
        "style_code",
        "supplier_name",
        "stock_days",
        "department_name",
        "category",
        "mrp",
    },
}

DAILY_UPDATE_SHEET = "daily_update"
DEFAULT_SHEET_ORDER = ["stores", "products", "inventory"]


def _import_models():
    for module_name in (
        "app.models.daily_snapshot",
        "app.models.delivery_logs",
        "app.models.inventory",
        "app.models.lifecycle",
        "app.models.product",
        "app.models.sales",
        "app.models.stores",
    ):
        importlib.import_module(module_name)


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _is_summary_value(value):
    if isinstance(value, str):
        value_text = value.strip().lower()
        if value_text and "total" in value_text:
            return True
    return False


def _is_footer_value(value):
    if not isinstance(value, str):
        return False
    value_text = value.strip().lower()
    return value_text.startswith("printed on") or value_text.startswith("generated on")


def _looks_like_header_row(row, header_set):
    matches = 0
    non_blank = 0
    for value in row:
        if _is_blank(value):
            continue
        non_blank += 1
        if normalize_header(value) in header_set:
            matches += 1
    return non_blank > 0 and matches == non_blank


def normalize_header(value):
    if value is None:
        return ""
    value_text = str(value).strip().lower()
    if not value_text:
        return ""
    for char in (" ", "-", ".", "/"):
        value_text = value_text.replace(char, "_")
    value_text = "_".join(part for part in value_text.split("_") if part)
    alias = HEADER_ALIASES.get(value_text)
    if alias:
        return alias
    alias = HEADER_ALIASES.get(value_text.replace("_", ""))
    if alias:
        return alias
    return value_text


def normalize_sheet_name(name):
    value_text = str(name).strip().lower()
    if not value_text:
        return ""
    for char in (" ", "-", ".", "/"):
        value_text = value_text.replace(char, "_")
    return "_".join(part for part in value_text.split("_") if part)


def normalize_sheet_list(value):
    if not value:
        return None
    if isinstance(value, (list, tuple, set)):
        items = [str(item).strip() for item in value if str(item).strip()]
    else:
        items = [part.strip() for part in str(value).split(",") if part.strip()]
    return items or None


def get_daily_update_aliases():
    settings = get_settings()
    aliases = normalize_sheet_list(settings.EXCEL_DAILY_UPDATE_SHEET_ALIASES)
    if not aliases:
        return []
    return [normalize_sheet_name(name) for name in aliases]


def apply_daily_update_aliases(sheet_map, aliases):
    if not aliases or DAILY_UPDATE_SHEET in sheet_map:
        return
    for alias in aliases:
        actual_name = sheet_map.get(alias)
        if actual_name:
            sheet_map[DAILY_UPDATE_SHEET] = actual_name
            break


def should_create_missing_stores():
    settings = get_settings()
    return bool(settings.EXCEL_CREATE_MISSING_STORES)


def ensure_store_exists(db, store_id):
    if not should_create_missing_stores():
        return
    if store_id is None:
        return
    if db.get(Store, store_id):
        return
    db.add(Store(id=store_id, name=f"Store {store_id}", city="Unknown"))
    db.flush()


def to_str(value, field, required=True):
    if _is_blank(value):
        if required:
            raise ValueError(f"{field} is required")
        return None
    return str(value).strip()


def to_int(value, field, required=True):
    if _is_blank(value):
        if required:
            raise ValueError(f"{field} is required")
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field} must be an integer")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"{field} must be an integer")
    if isinstance(value, str):
        value_text = value.strip()
        try:
            return int(value_text)
        except ValueError:
            try:
                numeric = float(value_text)
            except ValueError:
                raise ValueError(f"{field} must be an integer") from None
            if not numeric.is_integer():
                raise ValueError(f"{field} must be an integer")
            return int(numeric)
    return int(value)


def to_float(value, field):
    if _is_blank(value):
        raise ValueError(f"{field} is required")
    if isinstance(value, str):
        value = value.replace(",", "")
    return float(value)


def to_date(value, field):
    if _is_blank(value):
        raise ValueError(f"{field} is required")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value_text = value.strip()
        try:
            return date.fromisoformat(value_text)
        except ValueError:
            pass
        for fmt in ("%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value_text, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"{field} must be a date (YYYY-MM-DD)")


def load_sheet_rows(worksheet):
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return [], set()
    header_keys = [normalize_header(header) for header in headers]
    indices = [(idx, key) for idx, key in enumerate(header_keys) if key]
    columns = {key for key in header_keys if key}
    store_idx = None
    for idx, key in indices:
        if key == "store_id":
            store_idx = idx
            break

    rows = []
    for row in rows_iter:
        if row is None or all(_is_blank(value) for value in row):
            continue
        if _looks_like_header_row(row, columns):
            continue
        if store_idx is not None and store_idx < len(row):
            store_value = row[store_idx]
            if _is_blank(store_value):
                continue
            if _is_summary_value(store_value):
                continue
            if _is_footer_value(store_value):
                continue
        record = {key: row[idx] for idx, key in indices}
        rows.append(record)
    return rows, columns


def validate_columns(sheet_name, columns):
    required = REQUIRED_COLUMNS[sheet_name]
    missing = sorted(required - columns)
    if sheet_name == DAILY_UPDATE_SHEET and "stock_days" in missing:
        if "lifecycle_start_date" in columns:
            missing.remove("stock_days")
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(f"{sheet_name} sheet missing columns: {missing_text}")


def get_existing(db, model, record_id, *filters):
    if record_id:
        instance = db.get(model, record_id)
        if instance:
            return instance
    if filters:
        return db.execute(sa_select(model).where(*filters)).scalars().first()
    return None


def apply_upsert(db, instance, model, values):
    if instance:
        for key, value in values.items():
            setattr(instance, key, value)
        return "updated"
    db.add(model(**values))
    return "inserted"


def upsert_store(db, row):
    store_id = to_int(row.get("id"), "id", required=False)
    name = to_str(row.get("name"), "name")
    city = to_str(row.get("city"), "city")

    store = get_existing(db, Store, store_id, Store.name == name, Store.city == city)
    values = {"name": name, "city": city}
    return apply_upsert(db, store, Store, values)


def upsert_product(db, row):
    product_id = to_int(row.get("id"), "id", required=False)
    store_id = to_int(row.get("store_id"), "store_id")
    ensure_store_exists(db, store_id)
    style_code = to_str(row.get("style_code"), "style_code")
    barcode = to_str(row.get("barcode"), "barcode")
    article_name = to_str(row.get("article_name"), "article_name")
    category = to_str(row.get("category"), "category")
    supplier_name = to_str(row.get("supplier_name"), "supplier_name")
    mrp = to_float(row.get("mrp"), "mrp")
    department_value = row.get("department_name")
    department_name = None
    if not _is_blank(department_value):
        department_name = to_str(department_value, "department_name")

    product = get_existing(
        db,
        Product,
        product_id,
        Product.store_id == store_id,
        Product.style_code == style_code,
        Product.barcode == barcode,
    )
    values = {
        "store_id": store_id,
        "style_code": style_code,
        "barcode": barcode,
        "article_name": article_name,
        "category": category,
        "supplier_name": supplier_name,
        "mrp": mrp,
    }
    if department_name is not None:
        values["department_name"] = department_name
    elif not product:
        values["department_name"] = ""
    return apply_upsert(db, product, Product, values)


def upsert_inventory(db, row):
    inventory_id = to_int(row.get("id"), "id", required=False)
    store_id = to_int(row.get("store_id"), "store_id")
    ensure_store_exists(db, store_id)
    product_id = to_int(row.get("product_id"), "product_id")
    quantity = to_int(row.get("quantity"), "quantity")
    cost_price = to_float(row.get("cost_price"), "cost_price")
    current_price = to_float(row.get("current_price"), "current_price")
    lifecycle_start_date = to_date(row.get("lifecycle_start_date"), "lifecycle_start_date")

    inventory = get_existing(
        db,
        Inventory,
        inventory_id,
        Inventory.store_id == store_id,
        Inventory.product_id == product_id,
    )
    values = {
        "store_id": store_id,
        "product_id": product_id,
        "quantity": quantity,
        "cost_price": cost_price,
        "current_price": current_price,
        "lifecycle_start_date": lifecycle_start_date,
    }
    return apply_upsert(db, inventory, Inventory, values)


def resolve_price(value, fallback, field):
    if not _is_blank(value):
        return to_float(value, field)
    if _is_blank(fallback):
        raise ValueError(f"{field} is required")
    return to_float(fallback, field)


def resolve_lifecycle_start_date(row):
    lifecycle_value = row.get("lifecycle_start_date")
    if not _is_blank(lifecycle_value):
        return to_date(lifecycle_value, "lifecycle_start_date")
    stock_days = to_int(row.get("stock_days"), "stock_days")
    return date.today() - timedelta(days=stock_days)


def upsert_product_from_daily_update(db, row):
    store_id = to_int(row.get("store_id"), "store_id")
    ensure_store_exists(db, store_id)
    style_code = to_str(row.get("style_code"), "style_code")
    barcode_value = row.get("barcode")
    barcode = (
        to_str(barcode_value, "barcode")
        if not _is_blank(barcode_value)
        else style_code
    )
    article_value = row.get("article_name")
    article_name = (
        to_str(article_value, "article_name")
        if not _is_blank(article_value)
        else style_code
    )
    category = to_str(row.get("category"), "category")
    supplier_name = to_str(row.get("supplier_name"), "supplier_name")
    mrp = to_float(row.get("mrp"), "mrp")
    department_name = to_str(row.get("department_name"), "department_name")

    product = get_existing(
        db,
        Product,
        None,
        Product.store_id == store_id,
        Product.style_code == style_code,
        Product.barcode == barcode,
    )
    values = {
        "store_id": store_id,
        "style_code": style_code,
        "barcode": barcode,
        "article_name": article_name,
        "category": category,
        "supplier_name": supplier_name,
        "mrp": mrp,
        "department_name": department_name,
    }
    if product:
        for key, value in values.items():
            setattr(product, key, value)
        return "updated", product
    product = Product(**values)
    db.add(product)
    db.flush()
    return "inserted", product


def upsert_inventory_from_daily_update(db, row, product):
    store_id = product.store_id
    product_id = product.id
    quantity = to_int(row.get("quantity"), "quantity", required=False)
    if quantity is None:
        quantity = 0
    mrp = to_float(row.get("mrp"), "mrp")
    cost_price = resolve_price(row.get("cost_price"), mrp, "cost_price")
    current_price = resolve_price(row.get("current_price"), mrp, "current_price")
    lifecycle_start_date = resolve_lifecycle_start_date(row)

    inventory = get_existing(
        db,
        Inventory,
        None,
        Inventory.store_id == store_id,
        Inventory.product_id == product_id,
    )
    values = {
        "store_id": store_id,
        "product_id": product_id,
        "quantity": quantity,
        "cost_price": cost_price,
        "current_price": current_price,
        "lifecycle_start_date": lifecycle_start_date,
    }
    return apply_upsert(db, inventory, Inventory, values)


def import_daily_update_row(db, row):
    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    product_action, product = upsert_product_from_daily_update(db, row)
    counts[product_action] += 1
    inventory_action = upsert_inventory_from_daily_update(db, row, product)
    counts[inventory_action] += 1
    return counts


def import_rows(db, sheet_name, rows):
    counts = {"inserted": 0, "updated": 0, "skipped": 0}
    for row in rows:
        if sheet_name == DAILY_UPDATE_SHEET:
            row_counts = import_daily_update_row(db, row)
            for key, value in row_counts.items():
                counts[key] += value
            continue
        if sheet_name == "stores":
            action = upsert_store(db, row)
        elif sheet_name == "products":
            action = upsert_product(db, row)
        elif sheet_name == "inventory":
            action = upsert_inventory(db, row)
        else:
            raise ValueError(f"Unsupported sheet: {sheet_name}")
        counts[action] += 1
    return counts


def import_workbook(workbook_path, sheets=None, dry_run=False):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"File not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported.")

    workbook = load_workbook(workbook_path, data_only=True)
    sheet_map = {normalize_sheet_name(name): name for name in workbook.sheetnames}
    daily_update_aliases = get_daily_update_aliases()
    apply_daily_update_aliases(sheet_map, daily_update_aliases)

    sheet_list = normalize_sheet_list(sheets)
    if sheet_list:
        requested = []
        for name in sheet_list:
            key = normalize_sheet_name(name)
            if key in daily_update_aliases:
                key = DAILY_UPDATE_SHEET
            requested.append(key)
    else:
        if DAILY_UPDATE_SHEET in sheet_map:
            requested = [DAILY_UPDATE_SHEET]
        else:
            requested = [name for name in DEFAULT_SHEET_ORDER if name in sheet_map]

    if not requested:
        if len(sheet_map) == 1:
            only_key = next(iter(sheet_map))
            sheet_map[DAILY_UPDATE_SHEET] = sheet_map[only_key]
            requested = [DAILY_UPDATE_SHEET]
        else:
            raise ValueError("No matching sheets found to import.")

    _import_models()
    Base.metadata.create_all(bind=engine)
    ensure_sqlite_schema()

    results = {}
    db = SessionLocal()
    try:
        for sheet_key in requested:
            actual_name = sheet_map.get(sheet_key)
            if not actual_name:
                raise ValueError(f"Sheet not found: {sheet_key}")
            worksheet = workbook[actual_name]
            rows, columns = load_sheet_rows(worksheet)
            validate_columns(sheet_key, columns)
            results[sheet_key] = import_rows(db, sheet_key, rows)

        if dry_run:
            db.rollback()
        else:
            db.commit()
    except SQLAlchemyError:
        db.rollback()
        raise
    finally:
        db.close()

    return results

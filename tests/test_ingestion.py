import base64
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

from app.services import ingestion_service
from app.services.ingestion_service import (
    load_sheet_rows,
    normalize_header,
    normalize_sheet_list,
    validate_columns,
)


class IngestionServiceTest(unittest.TestCase):
    def test_header_aliases(self):
        self.assertEqual(normalize_header("Style Code"), "style_code")
        self.assertEqual(normalize_header("Item MRP"), "mrp")
        self.assertEqual(normalize_header("Lifecycle Start Date"), "lifecycle_start_date")

    def test_normalize_sheet_list(self):
        self.assertEqual(
            normalize_sheet_list("stores, products, inventory"),
            ["stores", "products", "inventory"],
        )

    def test_load_sheet_rows_supports_card_layout(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "001000000000760_20260220120225"

        worksheet.cell(row=6, column=7, value="Supplier Name")
        worksheet.cell(row=6, column=9, value="Sampat Sachin Sarees")
        worksheet.cell(row=8, column=7, value="Department")
        worksheet.cell(row=8, column=9, value="DRESS")
        worksheet.cell(row=9, column=7, value="Style")
        worksheet.cell(row=9, column=9, value="[None]")
        worksheet.cell(row=10, column=7, value="ItemCode")
        worksheet.cell(row=10, column=9, value="001261397")
        worksheet.cell(row=12, column=7, value="MRP")
        worksheet.cell(row=12, column=9, value="6995.00")
        worksheet.cell(row=13, column=7, value="Image Name")
        worksheet.cell(row=13, column=9, value="001261397")
        worksheet.cell(row=15, column=7, value="PUR Qty")
        worksheet.cell(row=15, column=9, value="1.00")

        rows, columns = load_sheet_rows(worksheet)

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["store_id"], 1)
        self.assertEqual(row["supplier_name"], "Sampat Sachin Sarees")
        self.assertEqual(row["style_code"], "001261397")
        self.assertEqual(row["category"], "DRESS")
        self.assertEqual(row["department_name"], "DRESS")
        self.assertEqual(row["stock_days"], 0)
        self.assertEqual(row["quantity"], "1.00")
        self.assertEqual(row["image_url"], "001261397")
        self.assertEqual(row["mrp"], "6995.00")
        validate_columns("daily_update", columns)

    def test_load_sheet_rows_extracts_embedded_images_from_card_layout(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "001000000000760_20260220120225"

        worksheet.cell(row=6, column=7, value="Supplier Name")
        worksheet.cell(row=6, column=9, value="Sampat Sachin Sarees")
        worksheet.cell(row=8, column=7, value="Department")
        worksheet.cell(row=8, column=9, value="DRESS")
        worksheet.cell(row=9, column=7, value="Style")
        worksheet.cell(row=9, column=9, value="PSD3-80279")
        worksheet.cell(row=10, column=7, value="ItemCode")
        worksheet.cell(row=10, column=9, value="001261397")
        worksheet.cell(row=12, column=7, value="MRP")
        worksheet.cell(row=12, column=9, value="6995.00")
        worksheet.cell(row=13, column=7, value="Image Name")
        worksheet.cell(row=13, column=9, value="001261397")
        worksheet.cell(row=15, column=7, value="PUR Qty")
        worksheet.cell(row=15, column=9, value="1.00")

        png_bytes = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/n1cAAAAASUVORK5CYII="
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image_path = tmp_path / "embedded.png"
            image_path.write_bytes(png_bytes)

            worksheet.add_image(XLImage(str(image_path)), "E5")
            workbook_path = tmp_path / "card_layout.xlsx"
            workbook.save(workbook_path)

            loaded_workbook = load_workbook(workbook_path, data_only=True)
            loaded_sheet = loaded_workbook[loaded_workbook.sheetnames[0]]
            image_dir = tmp_path / "images"

            with patch.object(ingestion_service, "_IMAGE_DIR", image_dir):
                rows, _ = load_sheet_rows(loaded_sheet)

            self.assertEqual(len(rows), 1)
            image_url = rows[0].get("image_url")
            self.assertIsNotNone(image_url)
            self.assertTrue(image_url.startswith("/static/images/001261397."))
            self.assertEqual(len(list(image_dir.glob("001261397.*"))), 1)

    def test_load_sheet_rows_prefers_embedded_image_over_identifier_text(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "daily_update"
        worksheet.append(
            [
                "store_id",
                "supplier_name",
                "stock_days",
                "style_code",
                "department_name",
                "category_name",
                "item_mrp",
                "cbs_qty",
                "image",
            ]
        )
        worksheet.append([1, "Supplier", 2, "STYLE-123", "DRESS", "DRESS", 4999, 1, "STYLE-123"])

        png_bytes = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/n1cAAAAASUVORK5CYII="
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image_path = tmp_path / "embedded.png"
            image_path.write_bytes(png_bytes)

            worksheet.add_image(XLImage(str(image_path)), "I2")
            workbook_path = tmp_path / "tabular_layout.xlsx"
            workbook.save(workbook_path)

            loaded_workbook = load_workbook(workbook_path, data_only=True)
            loaded_sheet = loaded_workbook[loaded_workbook.sheetnames[0]]
            image_dir = tmp_path / "images"

            with patch.object(ingestion_service, "_IMAGE_DIR", image_dir):
                rows, _ = load_sheet_rows(loaded_sheet)

            self.assertEqual(len(rows), 1)
            image_url = rows[0].get("image_url")
            self.assertIsNotNone(image_url)
            self.assertTrue(image_url.startswith("/static/images/STYLE-123."))
            self.assertEqual(len(list(image_dir.glob("STYLE-123.*"))), 1)

    def test_resolve_image_url_rejects_non_image_like_suffix(self):
        with patch.object(ingestion_service, "_get_image_index", return_value={}):
            image_url, image_explicit = ingestion_service.resolve_image_url({"image_url": "C.C[12]"})

        self.assertTrue(image_explicit)
        self.assertIsNone(image_url)

    def test_build_product_values_clears_invalid_explicit_image(self):
        existing = type("ProductStub", (), {"image_url": "/static/images/old.jpg"})()
        values = ingestion_service.build_product_values(
            store_id=1,
            style_code="S1",
            barcode="B1",
            article_name="A1",
            category="C1",
            supplier_name="SUP",
            mrp=100.0,
            product=existing,
            image_url=None,
            image_explicit=True,
        )

        self.assertIn("image_url", values)
        self.assertIsNone(values["image_url"])


if __name__ == "__main__":
    unittest.main()
